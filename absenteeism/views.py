from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from ImanSite.settings import THEMPLATE_NAME
from .models import Times,User
from datetime import datetime
import jdatetime
import json
from django.http import HttpResponse
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import numpy as np
from django.http.response import HttpResponseRedirect
from django.urls import reverse
import locale

# Create your views here.

@login_required
def index(request):
    context={}


    tm, status = Times.objects.get_or_create(
        date=datetime.now().date(),
        owner=request.user,
    )

    FMT = '%H:%M:%S'

    # this two lines use for get to day month and to day year
    monthNumber = jdatetime.date.today().month
    yearNumber = jdatetime.date.today().year

    # this two lines calculate start and end of month
    firstDayThisMonth = jdatetime.datetime(yearNumber, monthNumber, 1,00,00,00)
    if monthNumber == 12:
        firstDayNextMonth = jdatetime.datetime(yearNumber + 1, 1, 1,00,00,00)
    else:
        firstDayNextMonth = jdatetime.datetime(yearNumber,monthNumber + 1,1,00,00,00)

    # this two lines convert persian date to gregorian date and time
    firstDayThisMonth = firstDayThisMonth.togregorian()
    firstDayNextMonth = firstDayNextMonth.togregorian()

    # this two lines convert date time to date
    firstDayThisMonth = datetime.date(firstDayThisMonth)
    firstDayNextMonth = datetime.date(firstDayNextMonth)

    if request.method == 'POST':
        if request.POST.get('action',None) == "settime":
            times = tm.times
            times = json.loads(times)
            times.append(datetime.now().time().strftime("%H:%M:%S"))
            tm.times = json.dumps(times)
            if len(times)%2 == 1:
                del times[len(times) - 1]

            i = 0
            sum = '00:00:00'
            sum = datetime.strptime(sum, FMT)

            while i < len(times) / 2:
                s1 = times[i*2]
                s2 = times[i*2+1]
                tdelta = datetime.strptime(s2, FMT) - datetime.strptime(s1, FMT)
                i += 1
                sum += tdelta
            print('tm is:', tm.times, tm.date)
            tm.sum = sum
            tm.save()


        else:
            # this tow lines get date from filter buttons
            dateForFilter = request.POST.get("end_date", None)
            dateForFilterSplieted = dateForFilter.split("/")

            # this two lines use for get month and year of date that selected
            monthNumber = int(dateForFilterSplieted[1])
            yearNumber = int(dateForFilterSplieted[0])

            # this two lines calculate start and end of month
            firstDayThisMonth = jdatetime.datetime(yearNumber, monthNumber, 1, 00, 00, 00)
            firstDayNextMonth = jdatetime.datetime(yearNumber, monthNumber + 1, 1, 00, 00, 00)

            # this two lines convert persian date to gregorian date and time
            firstDayThisMonth = firstDayThisMonth.togregorian()
            firstDayNextMonth = firstDayNextMonth.togregorian()

            # this two lines convert date time to date
            firstDayThisMonth = datetime.date(firstDayThisMonth)
            firstDayNextMonth = datetime.date(firstDayNextMonth)





    # this line use for query to filter object base on needs
    context["time"]=Times.objects.filter(owner=request.user, date__range=[firstDayThisMonth,firstDayNextMonth])
    print('contest itmme', context['time'])
    context["sum"]=0

    # this loop use for calculate sum of times that user is needs
    for time in context["time"]:
        oddtime = {}
        context['sum']+=time.sum.hour*60*60+time.sum.minute*60+time.sum.second
        oddtime["stime"] = []
        oddtime['etime'] = []
        time.times = json.loads(time.times)
        i=0
        while i < len(time.times) / 2:
            try:
                oddtime["stime"].append(time.times[i*2])
                oddtime['etime'].append(time.times[i*2+1])
            except:
                pass
            i+=1
        time.times=oddtime

    return render(request,THEMPLATE_NAME+"/times.html", context)

@login_required
def adminT(request):

    context = {}

    context["user"]=[]
    user = User.objects.all()
    context["user"]=user

    # this two lines use for get to day month and to day year
    monthNumber = jdatetime.date.today().month
    yearNumber = jdatetime.date.today().year

    # this two lines calculate start and end of month
    firstDayThisMonth = jdatetime.datetime(yearNumber, monthNumber, 1,00,00,00)
    if monthNumber == 12:
        firstDayNextMonth = jdatetime.datetime(yearNumber + 1, 1, 1,00,00,00)
    else:
        firstDayNextMonth = jdatetime.datetime(yearNumber,monthNumber + 1,1,00,00,00)

    # this two lines convert persian date to gregorian date and time
    firstDayThisMonth = firstDayThisMonth.togregorian()
    firstDayNextMonth = firstDayNextMonth.togregorian()

    # this two lines convert date time to date
    firstDayThisMonth = datetime.date(firstDayThisMonth)
    firstDayNextMonth = datetime.date(firstDayNextMonth)


    if request.method == 'POST':
        # this tow lines get date from filter buttons
        dateForFilter = request.POST.get("end_date", None)
        dateForFilterSplieted = dateForFilter.split("/")

        # this two lines use for get month and year of date that selected
        monthNumber = int(dateForFilterSplieted[1])
        yearNumber = int(dateForFilterSplieted[0])

        # this two lines calculate start and end of month
        firstDayThisMonth = jdatetime.datetime(yearNumber, monthNumber, 1, 00, 00, 00)
        firstDayNextMonth = jdatetime.datetime(yearNumber, monthNumber + 1, 1, 00, 00, 00)

        # this two lines convert persian date to gregorian date and time
        firstDayThisMonth = firstDayThisMonth.togregorian()
        firstDayNextMonth = firstDayNextMonth.togregorian()

        # this two lines convert date time to date
        firstDayThisMonth = datetime.date(firstDayThisMonth)
        firstDayNextMonth = datetime.date(firstDayNextMonth)


    context["sum"] = 0

    # this line use for query to filter object base on needs
    owner = request.GET.get("id",1)
    ownerName = User.objects.get(id=owner)
    context["owners"] = ownerName
    request.session['selected_user'] = owner

    context["time"] = Times.objects.filter(owner=owner, date__range=[firstDayThisMonth, firstDayNextMonth])
    context["xlstime"]='%s-%s'%(yearNumber,monthNumber)
    # this two lines use for set month and year in export execl url
    context['year']=str(yearNumber)
    if monthNumber < 10:
        context['month'] = "0" + str(monthNumber)
    else:
        context['month']=str(monthNumber)

    # this loop use for calculate sum of times that user is needs
    for time in context["time"]:
        oddtime = {}
        context['sum'] += time.sum.hour * 60 * 60 + time.sum.minute * 60 + time.sum.second
        oddtime["stime"] = []
        oddtime['etime'] = []
        time.times = json.loads(time.times)
        i = 0
        while i < len(time.times) / 2:
            try:
                oddtime["stime"].append(time.times[i * 2])
                oddtime['etime'].append(time.times[i * 2 + 1])
            except:
                pass
            i += 1
        time.times = oddtime

    if request.user.has_perm("absenteeism.can_seetimes"):
        return render(request, THEMPLATE_NAME + "/admintime.html", context)

@login_required
def changeTime(request,id):

    context = {}

    # the below query use for get time objects from selected ID
    selectedTime = Times.objects.all().filter(id=id)
    timesObjects = []

    # the below loop use for convert time object to json array
    for time in selectedTime:
        timesObjects = time.times

    # the 'data' variable use for convert json array list
    data = json.loads(timesObjects)

    # the 'timesArray' variable user for convert list to array
    timesArray = np.asarray(data)

    # below loop and variables use for separate enter time and exit time
    context['sTime'] = []
    context['eTime'] = []
    context['allTime'] = {}
    pureTimes = {}
    pureTimes['sTime'] = []
    pureTimes['eTime'] = []
    t = 0

    while t < len(timesArray):
        context['sTime'].append(timesArray[t])
        pureTimes['sTime'].append(timesArray[t])
        t = t + 1
        if len(timesArray) > t:
            context['eTime'].append(timesArray[t])
            pureTimes['eTime'].append(timesArray[t])
            t = t + 1
        else:
            context['eTime'].append("00:00:00")
            pureTimes['eTime'].append("00:00:00")

    context['allTime'] = pureTimes

    editedTime = []
    arrayOfSumTimes = []
    FMT = '%H:%M:%S'
    if 'changeTime' in request.POST:
        for i in range(len(context['sTime'])):
            sTime = request.POST.get(str(i))
            editedTime.append(sTime)
            outNmae = "out" + str(i)
            eTime = request.POST.get(str(outNmae))
            editedTime.append(eTime)

            s1 = sTime
            s2 = eTime  # for example

            tdelta = datetime.strptime(s2, FMT) - datetime.strptime(s1, FMT)
            arrayOfSumTimes.append(tdelta)

        timeList = []
        for i in arrayOfSumTimes:
            timeList.append(str(i))

        totalSecs = 0
        for tm in timeList:
            timeParts = [int(s) for s in tm.split(':')]
            totalSecs += (timeParts[0] * 60 + timeParts[1]) * 60 + timeParts[2]
        totalSecs, sec = divmod(totalSecs, 60)
        hr, min = divmod(totalSecs, 60)
        timesss = str(hr) + ":" + str(min) + ":" + str(sec)
        sum = datetime.strptime(timesss, '%H:%M:%S').time()

        json_string = json.dumps(editedTime)

        Times.objects.filter(pk=id).update(times=json_string, sum=sum)
        return HttpResponseRedirect(reverse('changeTime', kwargs={'id': id}))

    elif 'newTime' in request.POST:

        # the two below lines use for append start and end time for new time
        context['sTime'].insert(len(context['sTime']), "00:00:00")
        context['eTime'].insert(len(context['eTime']), "00:00:00")

        # the below for loop use for put all of the start time and end time in a one array
        allTime = []
        t = 0
        for i in context['sTime']:
            print("s time is", i)
            allTime.append(i)
            allTime.append((context['eTime'])[t])
            t = t + 1

        # the below line use for convert array to json array of all times array
        jsonNewTime = json.dumps(allTime)

        # the below query use for update times
        Times.objects.filter(pk=id).update(times=jsonNewTime)

        return HttpResponseRedirect(reverse('changeTime', kwargs={'id': id}))

    elif "0" in request.POST:
        print('delete')

    return render(request, THEMPLATE_NAME + "/change user time.html", context)

@login_required
def vecation(request):
    if request.method == 'POST':
        dateOfDesc = request.POST.get('dateOfDesc')
        dateOfDescSplited = dateOfDesc.split("/")

        # this three lines use for get month and year of date that selected
        monthNumber = int(dateOfDescSplited[1])
        yearNumber = int(dateOfDescSplited[0])
        dayNumber = int(dateOfDescSplited[2])

        # this tow lines use for convert persian date to gregorian
        dateOfDesc = jdatetime.datetime(yearNumber, monthNumber, dayNumber)
        dateOfDesc = dateOfDesc.togregorian()

        desc = request.POST.get('desc')
        dateOfDesc = dateOfDesc.date()

        try:
            Times.objects.filter(owner=request.user,date=dateOfDesc).update(desc=desc)
        except:
            pass

    return render(request, THEMPLATE_NAME + "/vacation.html")


def export_users_xls(request,year,month):
    # this line use for query to filter object base on needs of users that in
    context = {}
    owener = request.session.get('selected_user')
    ownerName = User.objects.get(id=owener)

    # this two lines use for get to day month and to day year
    monthNumber = int(month)
    yearNumber = int(year)

    # this two lines calculate start and end of month
    firstDayThisMonth = jdatetime.datetime(yearNumber,monthNumber,1,00,00,00)
    firstDayNextMonth = jdatetime.datetime(yearNumber,monthNumber + 1,1,00,00,00)

    # this two lines convert persian date to gregorian date and time
    firstDayThisMonth = firstDayThisMonth.togregorian()
    firstDayNextMonth = firstDayNextMonth.togregorian()

    # this two lines convert date time to date
    firstDayThisMonth = datetime.date(firstDayThisMonth)
    firstDayNextMonth = datetime.date(firstDayNextMonth)

    row_queryset = Times.objects.filter(owner=owener, date__range=[firstDayThisMonth, firstDayNextMonth])

    # this loop use for calculate sum of times that user is needs
    sum = 0
    for time in row_queryset:
        sum += time.sum.hour * 60 * 60 + time.sum.minute * 60 + time.sum.second

    # the below lines use for calculate sum of times after update
    sec = str(sum % 60)
    min = str(sum//60%60)
    if len(sec) <2:
        sec = "0"+sec
    if len(min)<2:
        min = "0"+min

    # the below lines use for get cash of each selected user
    cash = User.objects.filter(id=owener).values('cash')
    for cashes in cash:
        context['cash'] = cashes['cash']

    # the below lines use for calculate cash of payment for selected user
    cashForHou = (sum // 3600) * context['cash']
    cashForMin = ((int(min) % 60 / 100) * context['cash'])
    payCash = cashForHou + cashForMin
    locale.setlocale(locale.LC_ALL, '')
    payCash = '{0:,}'.format(payCash).replace('.0', '')

    # the below line use for convert  time to  regular time for show in excel file
    sum = str(sum // 3600) + ":" + min + ":" + sec

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="{}-times.xls"'.format(ownerName)
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Times'

    # Define the titles for columns
    columns = [
        'تاریخ',
        'ساعت ورود و خروج',
        'جمع ساعت',
        'توضیحات',
    ]

    row_num = 1

    border_bottom = Border(
        bottom=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
    )

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = border_bottom
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(name='B Nazanin', bold=True, size=13)

    # Iterate through all movies
    for rows in row_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            str(jdatetime.date.fromgregorian(date=rows.date)),
            rows.times,
            rows.sum,
            rows.desc,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_bottom
            cell.font = Font(name='B Nazanin', bold=False, size=10)

    # the below lines use for set name and sum of the times in the footer of execl file
    cell = worksheet.cell(row=row_num + 1, column=1)
    cell.value = str(ownerName)
    cell.border = border_bottom
    cell.font = Font(name='B Nazanin', bold=False, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = worksheet.cell(row=row_num + 1, column=2)
    cell.value = str(payCash) + 'ریال'
    cell.border = border_bottom
    cell.font = Font(name='B Nazanin', bold=False, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = worksheet.cell(row=row_num + 1, column=3)
    cell.value = sum
    cell.border = border_bottom
    cell.font = Font(name='B Nazanin', bold=False, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = worksheet.cell(row=row_num + 1, column=4)
    cell.value = ""
    cell.border = border_bottom
    cell.font = Font(name='B Nazanin', bold=False, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(response)
    return response
